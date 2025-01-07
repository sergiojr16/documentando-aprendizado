import openpyxl
from collections import Counter
from openpyxl.chart import BarChart, Reference



def identifica_indices_colunas(sheet):
    header = [celula.value for celula in next(sheet.iter_rows(max_row=1))]
    col_index_ref = header.index("Referência")
    col_index_item = header.index("Item")
    col_index_onde = header.index("Onde")
    col_index_valor = header.index("Valor")
    return col_index_item, col_index_valor, col_index_onde, col_index_ref


def define_filtro(possiveis_referencias):
    while True:
        print("Intervalo válido de referência: JANEIRO/2023 até JANEIRO/2025")
        filtro = input("\nDigite a referência (exemplo: JANEIRO/2024): ").upper()
        if filtro not in possiveis_referencias:
            print("\nPor favor, digite uma referência válida.")
        else:
            print(f"\nAnalisando {filtro}...")
            return filtro

def valida_referencias(sheet, col_index_ref):
    possiveis_referencias = set()
    for linha in sheet.iter_rows(min_row=2, values_only=True):
        ref = linha[col_index_ref]
        possiveis_referencias.add(ref)
    return possiveis_referencias


def analisa_historico_valores_mensais(sheet, col_index_ref, col_index_valor):
    valor_mensal = {}
    for linha in sheet.iter_rows(min_row=2, values_only=True):
        ref = linha[col_index_ref]
        valor = linha[col_index_valor]
        if isinstance(valor, (int, float)):
            if ref not in valor_mensal:
                valor_mensal[ref] = 0
            valor_mensal[ref] += valor
    for ref, valor in valor_mensal.items():
        print(f"No período de {ref} foi obtido {valor}kk em loots raros. ")
    return valor_mensal


def analisa_valores_mensais(filtro, sheet, col_index_ref, col_index_valor):
    valor_mensal = {}
    for linha in sheet.iter_rows(min_row=2, values_only=True):
        ref = linha[col_index_ref]
        valor = linha[col_index_valor]
        if ref == filtro:
            if isinstance(valor, (int, float)):
                if ref not in valor_mensal:
                    valor_mensal[ref] = 0
                valor_mensal[ref] += valor
    for ref, valor in valor_mensal.items():
        print(f"No período de {ref} foi obtido {valor}kk em loots raros. ")



def analisa_origem_valores(filtro, sheet, col_index_ref, col_index_onde, col_index_valor):
    origem_valores = {}

    for linha in sheet.iter_rows(min_row=2, values_only=True):
        ref = linha[col_index_ref]
        origem = linha[col_index_onde]
        valor = linha[col_index_valor]
        if ref == filtro:
            if isinstance(valor, (int, float)):
                if origem not in origem_valores:
                    origem_valores[origem] = 0
                origem_valores[origem] += valor

    print(f"\nNo período de {filtro} foram obtidos: \n")
    for origem, valor in origem_valores.items():
        print(f" - {valor}kks em loots raros vindos de {origem}. ")


def analisa_quantidade_items(filtro, sheet, col_index_ref, col_index_item):
    quantidade_items = {}
    contador = Counter()

    for linha in sheet.iter_rows(min_row=2, values_only=True):
        ref = linha[col_index_ref]
        item = linha[col_index_item]
        if ref == filtro:
            contador[item] += 1

    print(f"\nDurante o período de referência {filtro} foram obtidos: \n")
    for item, contagem in contador.items():
        quantidade_items[item] = contagem
        print(f" - {contagem}x {item}.")


def criar_novo_arquivo(valor_mensal):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Resultados Mensais - Rares"

    new_sheet.append(["Período Referência", "Valor Total"])
    for ref, valor in valor_mensal.items():
        new_sheet.append([ref, valor])

    grafico = BarChart()
    grafico.title = "Resultado Mensal"
    grafico.x_axis.title = "Período Referência"
    grafico.y_axis.title = "Valor Total"

    data = Reference(new_sheet, min_col=2, min_row=1, max_col=2, max_row=len(valor_mensal) + 1)
    categories = Reference(new_sheet, min_col=1, min_row=2, max_row=len(valor_mensal) + 1)
    grafico.add_data(data, titles_from_data=True)
    grafico.set_categories(categories)

    new_sheet.add_chart(grafico, "D2")

    new_file_path = "valores_mensais_rares.xlsx"
    try:
        new_workbook.save(new_file_path)
        print(f"\nNovo arquivo salvo com sucesso em {new_file_path} ")
    except Exception:
        print(f"Ocorreu um erro ao salvar o arquivo: {Exception}")
    new_workbook.close()


def menu_inicial():
    print("\n" + "-" * 40)
    print("==== AUTOMAÇÃO ANÁLISE - LOOTS ====")
    print("-" * 40)
    print("[1] - Analisar histórico valores mensais ")
    print("[2] - Selecionar mês para análise mensal ")
    print("[3] - Finalizar programa ")
    print("-" * 40)


def menu_mensal():
    print("\n" + "-" * 40)
    print("[1] - Analisar valores ")
    print("[2] - Analisar origem dos loots ")
    print("[3] - Analisar quantidade de cada loot ")
    print("[4] - Voltar ao menu anterior ")
    print("-" * 40)


def automacao_leitura_rares():
    arquivo = "loot_rares.xlsx"
    try:
        workbook = openpyxl.load_workbook(arquivo)
        print(f"O Arquivo {arquivo} foi aberto com sucesso. ")
    except Exception:
        print(f"Ocorreu um erro ao abrir o arquivo {Exception}")
        exit()
    sheet = workbook.active

    col_index_item, col_index_valor, col_index_onde, col_index_ref = identifica_indices_colunas(sheet)
    possiveis_referencias = valida_referencias(sheet, col_index_ref)
    while True:
        menu_inicial()
        opcao = input("Selecione uma opção: ")
        if opcao == "1":
            valor_mensal = analisa_historico_valores_mensais(sheet, col_index_ref, col_index_valor)
            opcao_novo_arquivo = input("Deseja criar um arquivo com a síntese de valores X mês? [S]/[N]").upper()
            if opcao_novo_arquivo == "S":
                print("Criando arquivo...")
                criar_novo_arquivo(valor_mensal)
            elif opcao_novo_arquivo == "N":
                return
            else:
                print("Não foi digitada uma opção válida.")
        elif opcao == "2":
            filtro = define_filtro(possiveis_referencias)
            while True:
                menu_mensal()
                opcao2 = input("Selecione uma opção: ")
                if opcao2 == "1":
                    analisa_valores_mensais(filtro, sheet, col_index_ref, col_index_valor)
                elif opcao2 == "2":
                    analisa_origem_valores(filtro, sheet, col_index_ref, col_index_onde, col_index_valor)
                elif opcao2 == "3":
                    analisa_quantidade_items(filtro, sheet, col_index_ref, col_index_item)
                elif opcao2 == "4":
                    break
                else:
                    print("Opção inválida, tente novamente. ")
        elif opcao == "3":
            print("Finalizando programa...")
            break
        else:
            print("Opção inválida, tente novamente. ")

    workbook.close()
    pass


if __name__ == "__main__":
    automacao_leitura_rares()
